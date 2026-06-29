import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import holidays
from streamlit_calendar import calendar
from supabase import create_client, Client
from ortools.sat.python import cp_model
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
import re
import io
import time
import os
import streamlit.components.v1 as components

# ------------------------------------------------------------------
# 1. ตั้งค่าหน้าเว็บ & เวทมนตร์ CSS (Modern UI)
# ------------------------------------------------------------------
st.set_page_config(page_title="PharmSuk App", layout="wide", page_icon="💊")

components.html(
    """
    <script>
    setInterval(function() {
        window.parent.document.dispatchEvent(new Event('mousemove'));
    }, 240000);
    </script>
    """,
    height=0, width=0
)

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Kanit:wght@300;400;500;600&display=swap');
    
    html, body, [class*="css"], .stMarkdown, .stText, p, h1, h2, h3, h4, h5, h6 {
        font-family: 'Kanit', sans-serif !important;
    }
    .block-container { padding-top: 1.5rem !important; padding-bottom: 2rem !important; }
    .stButton>button, div[data-testid="stFormSubmitButton"] button { border-radius: 8px !important; transition: all 0.2s ease !important; }
    .stButton>button:hover, div[data-testid="stFormSubmitButton"] button:hover { transform: translateY(-2px) !important; box-shadow: 0 4px 10px rgba(0,0,0,0.1) !important; }
    div[data-testid="stExpander"] { border: 1px solid rgba(240,242,246,0.5) !important; box-shadow: 0 2px 8px rgba(0,0,0,0.05) !important; border-radius: 10px !important; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] { border-radius: 8px 8px 0px 0px; padding: 10px 16px; background-color: #f8f9fa; }
    .stTabs [data-baseweb="tab"] p { color: #154360 !important; font-weight: 500 !important; }
    .stTabs [aria-selected="true"] { background-color: #E8DAEF !important; border-bottom: 3px solid #9B59B6 !important; }
    .stTabs [aria-selected="true"] p { color: #4A235A !important; font-weight: 600 !important; }
    div[data-testid="stMetricValue"] { font-size: 2rem !important; color: #9B59B6 !important; font-weight: 600 !important; }
    [data-testid="stSidebar"] { background: linear-gradient(180deg, rgba(21,67,96,1) 0%, rgba(77,208,225,1) 100%); color: white !important; }
    [data-testid="stSidebar"] .stMarkdown p, [data-testid="stSidebar"] .stRadio label p { color: rgba(255,255,255,0.9) !important; }
    [data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.1) !important; }
    [data-testid="stSidebar"] [data-baseweb="radio"] > div:first-child { display: none !important; }
    [data-testid="stSidebar"] div[role="radiogroup"] label { padding: 12px 15px; border-radius: 8px; margin-bottom: 5px; transition: all 0.2s ease; cursor: pointer; width: 100%; background-color: rgba(255,255,255,0.05) !important; border: 1px solid transparent !important; }
    [data-testid="stSidebar"] div[role="radiogroup"] label:hover { background-color: rgba(255,255,255,0.1) !important; }
    [data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) { background-color: rgba(255,255,255,0.25) !important; border-left: 5px solid #ffffff !important; border-radius: 4px 8px 8px 4px !important; }
    [data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) p { color: white !important; font-weight: 600 !important; }
    [data-testid="stSidebar"] .stButton > button { background-color: transparent !important; color: rgba(255,255,255,0.8) !important; border: 1px solid rgba(255,255,255,0.3) !important; border-radius: 8px !important; font-weight: 400 !important; box-shadow: none !important; }
    [data-testid="stSidebar"] .stButton > button:hover { background-color: rgba(255,255,255,0.1) !important; color: #ffffff !important; border-color: #ffffff !important; }
    [data-testid="stSidebar"] img { border-radius: 10px; box-shadow: 0 4px 15px rgba(0,0,0,0.3); margin-bottom: 15px; }
    </style>
""", unsafe_allow_html=True)

@st.cache_resource
def init_connection():
    raw_url = st.secrets["supabase"]["url"]
    raw_key = st.secrets["supabase"]["key"]
    return create_client(raw_url.strip().rstrip('/'), raw_key.strip())

try: supabase: Client = init_connection()
except Exception: supabase = None

# ------------------------------------------------------------------
# Global Variables
# ------------------------------------------------------------------
VALID_TIMES = ["08.30", "09.00", "09.30", "10.00", "10.30", "11.00", "11.30", "12.00", "12.30", "13.00", "13.30", "14.00", "14.30", "15.00", "15.30", "16.00", "16.30"]
time_slots = [f"{VALID_TIMES[i]}-{VALID_TIMES[i+1]}" for i in range(16)]
current_year = datetime.now().year
th_holidays = holidays.Thailand(years=[current_year - 1, current_year, current_year + 1])

if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.current_user = None

if 'auth_mode' not in st.session_state:
    st.session_state.auth_mode = 'login'

if "gen_df" not in st.session_state: st.session_state.gen_df = None
if "gen_html" not in st.session_state: st.session_state.gen_html = None
if "gen_excel" not in st.session_state: st.session_state.gen_excel = None
if "show_balloons" not in st.session_state: st.session_state.show_balloons = False
if 'pt_daily_db' not in st.session_state: st.session_state.pt_daily_db = [] 

def safe_idx(lst, val, default=0):
    try: return lst.index(val)
    except ValueError: return default

def get_thai_date(date_obj):
    if isinstance(date_obj, str): date_obj = datetime.strptime(date_obj, "%Y-%m-%d").date()
    thai_months = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
    thai_days = ["วันจันทร์", "วันอังคาร", "วันพุธ", "วันพฤหัสบดี", "วันศุกร์", "วันเสาร์", "วันอาทิตย์"]
    return f"{thai_days[date_obj.weekday()]}ที่ {date_obj.day} {thai_months[date_obj.month]} {date_obj.year + 543}"

# ------------------------------------------------------------------
# Database Functions
# ------------------------------------------------------------------
def fetch_users():
    if supabase:
        res = supabase.table("users").select("*").execute()
        return {user['username']: user for user in res.data}
    return {}

def fetch_requests():
    if supabase:
        res = supabase.table("requests").select("*").order("created_at", desc=True).execute()
        return res.data
    return []

def add_request(user_name, req_type, req_date, detail):
    if supabase:
        is_leave = "ลางาน" in req_type
        status = "⏳ รออนุมัติ" if is_leave else "✅ อนุมัติแล้ว"
        if is_leave and "ลาป่วย" in detail: status = "✅ อนุมัติแล้ว" 
        supabase.table("requests").insert({"user_name": user_name, "req_type": req_type, "req_date": req_date.strftime("%Y-%m-%d"), "detail": detail, "status": status}).execute()

def update_request_status(req_id, new_status):
    if supabase: supabase.table("requests").update({"status": new_status}).eq("id", req_id).execute()

def delete_request(req_id):
    if supabase: supabase.table("requests").delete().eq("id", req_id).execute()

def save_schedule_to_db(target_date_str, html_table):
    if supabase:
        try:
            res = supabase.table("schedules").select("id").eq("schedule_date", target_date_str).execute()
            data = {"schedule_date": target_date_str, "html_content": html_table, "created_at": datetime.now().isoformat()}
            if len(res.data) > 0: supabase.table("schedules").update(data).eq("schedule_date", target_date_str).execute()
            else: supabase.table("schedules").insert(data).execute()
            return True
        except: return False
    return False

def add_user_db(username, password, full_name, role, real_name="", surname="", email="", position="", display_order=99):
    if supabase:
        data = {"username": username, "password": password, "full_name": full_name, "role": role, "real_name": real_name, "surname": surname, "email": email, "position": position, "display_order": display_order}
        supabase.table("users").insert(data).execute()

def update_user_role(username, role):
    if supabase: supabase.table("users").update({"role": role}).eq("username", username).execute()

def update_user_order(username, new_order):
    if supabase: supabase.table("users").update({"display_order": new_order}).eq("username", username).execute()

def delete_user_db(username):
    if supabase: supabase.table("users").delete().eq("username", username).execute()

def update_user_profile(username, real_name, surname, email, position):
    if supabase:
        data = {"real_name": real_name, "surname": surname, "email": email, "position": position}
        supabase.table("users").update(data).eq("username", username).execute()

def update_user_password(username, new_password):
    if supabase: supabase.table("users").update({"password": new_password}).eq("username", username).execute()

def update_user_fullname(username, new_fullname):
    if supabase: supabase.table("users").update({"full_name": new_fullname}).eq("username", username).execute()

# ------------------------------------------------------------------
# Login & Auth System
# ------------------------------------------------------------------
def login_page():
    col_img1, col_img2, col_img3 = st.columns([1, 2, 1])
    with col_img2:
        if os.path.exists("banner.png"):
            st.markdown("""<style>[data-testid="stImage"] > img { border-radius: 0px !important; box-shadow: none !important; background-color: transparent !important; }</style>""", unsafe_allow_html=True)
            st.image("banner.png", use_container_width=True)
        else: st.markdown("<h1 style='text-align: center; color: #154360;'>💊 PharmSuk</h1>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.session_state.get('auth_mode', 'login') == 'login':
            with st.form("login_form"):
                st.markdown("<h3 style='text-align: center; margin-top:0;'>เข้าสู่ระบบ</h3>", unsafe_allow_html=True)
                login_val = st.text_input("Username หรือ Email")
                password = st.text_input("Password", type="password")
                submit_btn = st.form_submit_button("เข้าสู่ระบบ", type="primary", use_container_width=True)
                forgot_btn = st.form_submit_button("ลืมรหัสผ่าน?", type="secondary", use_container_width=True)
                
                if submit_btn:
                    user_input = login_val.lower().strip()
                    matched_user = None
                    users_db = fetch_users()
                    if user_input in users_db: matched_user = users_db[user_input]
                    else:
                        for u in users_db.values():
                            if u.get('email') and u.get('email').lower().strip() == user_input and user_input != "":
                                matched_user = u; break
                    if matched_user and matched_user['password'] == password:
                        st.session_state.logged_in = True
                        st.session_state.current_user = matched_user
                        st.rerun()
                    else: st.error("❌ Username/Email หรือ Password ไม่ถูกต้อง")
                if forgot_btn:
                    st.session_state.auth_mode = 'forgot'
                    st.rerun()
        else:
            with st.form("forgot_form"):
                st.markdown("<h3 style='text-align: center; margin-top:0;'>กู้คืนรหัสผ่าน</h3>", unsafe_allow_html=True)
                f_user = st.text_input("Username")
                f_email = st.text_input("อีเมล (Email) ที่ลงทะเบียนไว้")
                new_pass = st.text_input("ตั้งรหัสผ่านใหม่", type="password")
                reset_btn = st.form_submit_button("รีเซ็ตรหัสผ่าน", type="primary", use_container_width=True)
                back_btn = st.form_submit_button("กลับไปหน้าเข้าสู่ระบบ", type="secondary", use_container_width=True)
                
                if reset_btn:
                    user_clean = f_user.lower().strip()
                    users_db = fetch_users()
                    if user_clean in users_db and users_db[user_clean].get('email') == f_email and f_email != "":
                        update_user_password(user_clean, new_pass)
                        st.success("✅ เปลี่ยนรหัสผ่านสำเร็จ! กรุณาเข้าสู่ระบบใหม่")
                        time.sleep(2)
                        st.session_state.auth_mode = 'login'
                        st.rerun()
                    else: st.error("❌ Username หรือ Email ไม่ถูกต้อง หรือยังไม่ได้ผูกอีเมลไว้")
                if back_btn:
                    st.session_state.auth_mode = 'login'
                    st.rerun()

def logout():
    st.session_state.logged_in = False
    st.session_state.current_user = None

if not st.session_state.logged_in:
    login_page()
    st.stop()

user_info = st.session_state.current_user
if user_info is None:
    st.session_state.logged_in = False
    st.rerun()

def get_std_pos(u):
    p = u.get('position')
    return p if p in ["เภสัชกร", "ผู้ช่วยเภสัชกร"] else "เภสัชกร"

# ------------------------------------------------------------------
# Sidebar & Menu
# ------------------------------------------------------------------
with st.sidebar:
    if os.path.exists("banner.png"): st.image("banner.png", use_container_width=True)
    real_name = (user_info.get('real_name') or "").strip()
    surname = (user_info.get('surname') or "").strip()
    display_name = f"{real_name} {surname}".strip() if (real_name or surname) else user_info.get('full_name', '')
    
    st.markdown(f"### 👤 คุณ {display_name} ({user_info['role']})")
    st.button("🚪 ออกจากระบบ", on_click=logout, use_container_width=True)
    st.markdown("---")
    
    if user_info['role'] in ['Admin', 'Head']:
        st.session_state.current_view = st.radio("🔄 สลับมุมมองระบบ", ["เภสัชกร", "ผู้ช่วยเภสัชกร"], horizontal=True)
    else:
        st.session_state.current_view = get_std_pos(user_info)
        st.info(f"📍 มุมมอง: {st.session_state.current_view}")
    st.markdown("---")
    
    menu_options = ["🗓️ ปฏิทินห้องยา", "✏️ ลงข้อมูล & จัดการข้อมูล", "👤 ข้อมูลส่วนตัว"]
    if user_info['role'] in ['Admin', 'Head']: menu_options.extend(["🔐 อนุมัติคำขอ (Approve)", "📝 สร้างตารางทำงานประจำวัน", "🏃 จัดการพาร์ทไทม์", "👥 จัดการผู้ใช้งาน"])
    page = st.radio("เลือกเมนู", menu_options, label_visibility="collapsed")
    st.markdown("<br><hr style='margin:0; border-color: rgba(255,255,255,0.1);'><p style='text-align:center; color:rgba(255,255,255,0.4); font-size:12px; margin-top:5px;'>💡 PharmSuk v47</p>", unsafe_allow_html=True)

# ------------------------------------------------------------------
# Global Context & Queue Logic
# ------------------------------------------------------------------
current_view = st.session_state.current_view
users_db = fetch_users()
fullname_to_pos = {u['full_name']: get_std_pos(u) for u in users_db.values()}

active_users = [u for u in users_db.values() if u.get('role') != 'System' and get_std_pos(u) == current_view]
active_users.sort(key=lambda x: (x.get('display_order') if x.get('display_order') is not None else 99, x.get('full_name', '')))
base_pharmacist_list = [u['full_name'] for u in active_users]
head_pharmacist_list = [u['full_name'] for u in active_users if u.get('role') == 'Head']

def is_req_in_view(req):
    if req['user_name'] == "SYSTEM_REQ": return True
    return fullname_to_pos.get(req['user_name'], 'เภสัชกร') == current_view

raw_requests = fetch_requests()
all_requests = [r for r in raw_requests if is_req_in_view(r)]

# ==================================================================
# คำนวณคิวใหม่เสมอ! (ข้ามรายการที่ถูก "ยกเลิก" หรือ "ไม่อนุมัติ")
# ==================================================================
leave_queues = {}
leaves_by_date = {}
active_leaves_for_queue = [
    r for r in raw_requests 
    if "ลางาน" in r['req_type'] 
    and r['status'] in ["⏳ รออนุมัติ", "✅ อนุมัติแล้ว"] 
    and ("ลาพักร้อน" in r['detail'] or "ลากิจ" in r['detail'])
]
sorted_for_queue = sorted(active_leaves_for_queue, key=lambda x: x.get('created_at', ''))

for r in sorted_for_queue:
    dk = r['req_date']
    leaves_by_date[dk] = leaves_by_date.get(dk, 0) + 1
    leave_queues[r['id']] = leaves_by_date[dk]

# ------------------------------------------------------------------
# AI Helpers & UI Renderers
# ------------------------------------------------------------------
def load_db_to_dashboard(approved_today, all_requests, target_date_str):
    leaves, tasks, shifts, bo_list = [], [], [], []
    target_dt = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    for r in approved_today:
        if "ลางาน" in r['req_type']:
            l_type = "ทั้งวัน"
            if "ครึ่งวันเช้า" in r['detail']: l_type = "เช้า"
            elif "ครึ่งวันบ่าย" in r['detail']: l_type = "บ่าย"
            elif "ลาป่วย" in r['detail']: l_type = "ฉุกเฉิน"
            s_time, e_time = "08.30", "16.30"
            times = re.findall(r'\d{2}\.\d{2}', r['detail'])
            if len(times) >= 2: s_time, e_time = times[0], times[1]
            if l_type == "ฉุกเฉิน": leaves.append({"user_name": r['user_name'], "leave_type": (s_time, e_time), "start": s_time, "end": e_time, "detail": r['detail']})
            else: leaves.append({"user_name": r['user_name'], "leave_type": l_type, "start": s_time, "end": e_time, "detail": r['detail']})
        elif "งานพิเศษ" in r['req_type']:
            times = re.findall(r'\d{2}\.\d{2}', r['detail'])
            tasks.append({"user_name": r['user_name'], "task_name": r['detail'].split('(')[0].replace('งานพิเศษ:', '').strip(), "start": times[0] if len(times)>0 else "08.30", "end": times[1] if len(times)>1 else "16.30"})
        elif "ออกเวร" in r['req_type']:
            s_type = "ออกเวรดึก" if "ดึก" in r['detail'] else "ออกเวรเย็น"
            room = "ชั้น 1"
            if "พระเทพ" in r['detail']: room = "ตึกพระเทพ"
            elif "ตึกเก่า" in r['detail']: room = "ตึกเก่า"
            shifts.append({"user_name": r['user_name'], "shift_type": s_type, "room": room, "start": "08.30", "end": "10.30", "time_slot": "15.00-15.30"})
    for r in all_requests:
        if r['req_type'] == "Back Office" and r['status'] == "✅ อนุมัติแล้ว":
            try:
                start_dt = datetime.strptime(r['req_date'], "%Y-%m-%d").date()
                parts = r['detail'].split('|') 
                if len(parts) >= 4:
                    end_dt = datetime.strptime(parts[1], "%Y-%m-%d").date()
                    if start_dt <= target_dt <= end_dt:
                        if len(parts) >= 5:
                            thai_days = ["จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"]
                            if thai_days[target_dt.weekday()] not in parts[4].split(','): continue
                        times = parts[2].split('-')
                        bo_list.append({"user_name": r['user_name'], "task_name": parts[3], "start": times[0], "end": times[1] if len(times)>1 else "16.30"})
            except: pass
    return leaves, tasks, shifts, bo_list

def force_sync_dashboard(target_date_str, dashboard_requests):
    approved_today = [r for r in dashboard_requests if r["req_date"] == target_date_str and r["status"] == "✅ อนุมัติแล้ว"]
    pts_today = [pt for pt in st.session_state.pt_daily_db if pt["date"] == target_date_str and pt.get('position', 'เภสัชกร') == current_view]
    leaves, tasks, shifts, bo_list = load_db_to_dashboard(approved_today, dashboard_requests, target_date_str)
    
    st.session_state.dash_leaves = leaves
    st.session_state.dash_tasks = tasks
    st.session_state.dash_shifts = shifts
    st.session_state.dash_pts = pts_today
    st.session_state.dash_subs = [] 
    st.session_state.dash_locks = []
    st.session_state.dash_bo = bo_list
    st.session_state.dash_date = target_date_str
    st.session_state.dash_hash = len(dashboard_requests) + len(pts_today)

# --- AI Core Logic (For Pharmacists) ---
def time_to_slot(t_str): return safe_idx(VALID_TIMES, t_str, 0)
dispensing_tasks = [f"จ่าย {i}" for i in range(4, 12)]
ver_cpoe_tasks = ["Ver 1 INC", "Ver 2/ปณ.", "Ver 3/A", "Ver 4", "Ver 5", "Ver 6", "Ver 7", "Ver 8", "Ver 9", "Ver 10"]
ver_ps_tasks = [f"Ver PS{i}" for i in range(1, 11)]
base_main_tasks = dispensing_tasks + ver_cpoe_tasks + ver_ps_tasks + ["Match + C", "Match + C2"]

def generate_schedule(DAY_OF_WEEK, LEAVES, CUSTOM_TASKS, PART_TIME, FIX_BREAKS, FIXED_MAIN_TASKS, SICK_PEOPLE, IS_MWF, HEAD_PHARMACISTS, ALLOW_HEAD_ASSIST=False):
    ft_pharmacists = base_pharmacist_list
    head_pharmacists = HEAD_PHARMACISTS 
    pt_pharmacists = [pt['name'] for pt in PART_TIME]
    all_pharmacists = ft_pharmacists + pt_pharmacists
    error_msgs = []
    leave_slots_check = {}
    
    for p in ft_pharmacists:
        if p in LEAVES:
            l_type = LEAVES[p]
            if l_type == 'ทั้งวัน': l_range = range(0, 16)
            elif l_type == 'เช้า': l_range = range(0, 9)
            elif l_type == 'บ่าย': l_range = range(7, 16)
            else: l_range = range(time_to_slot(l_type[0]), time_to_slot(l_type[1]))
            for t in l_range: leave_slots_check[(p, t)] = l_type

    custom_slots_check = {}
    for (p, start, end), t_name in CUSTOM_TASKS.items():
        s_idx, e_idx = time_to_slot(start), time_to_slot(end)
        for t in range(s_idx, e_idx):
            if (p, t) in leave_slots_check: error_msgs.append(f"⚠️ **{p}**: ถูกตั้งให้ **ลา** และทำภารกิจ **{t_name}** ทับซ้อนกันเวลา {time_slots[t]}")
            if (p, t) in custom_slots_check: error_msgs.append(f"⚠️ **{p}**: ถูกตั้งภารกิจ **{custom_slots_check[(p, t)]}** และ **{t_name}** ทับซ้อนกันเวลา {time_slots[t]}")
            custom_slots_check[(p, t)] = t_name

    fixed_slots_check = {}
    for (p, start, end), t_name in FIXED_MAIN_TASKS.items():
        s_idx, e_idx = time_to_slot(start), time_to_slot(end)
        for t in range(s_idx, e_idx):
            if (p, t) in leave_slots_check: error_msgs.append(f"⚠️ **{p}**: ลา และล็อกงานหลัก **{t_name}** ทับซ้อนกันเวลา {time_slots[t]}")
            if (p, t) in custom_slots_check: error_msgs.append(f"⚠️ **{p}**: มีภารกิจ **{custom_slots_check[(p, t)]}** และล็อกงานหลัก **{t_name}** ทับซ้อนกันเวลา {time_slots[t]}")
            if (p, t) in fixed_slots_check: error_msgs.append(f"⚠️ **{p}**: ถูกล็อกงานหลัก **{fixed_slots_check[(p, t)]}** และ **{t_name}** ทับซ้อนกันเวลา {time_slots[t]}")
            fixed_slots_check[(p, t)] = t_name

    if error_msgs:
        return None, "Validation Failed", "ตรวจพบปัญหาในการตั้งค่า กรุณาแก้ไข:\n\n" + "\n".join(list(dict.fromkeys(error_msgs))[:10])

    model = cp_model.CpModel()
    tasks = base_main_tasks + ['Matching', 'พัก', 'งานเฉพาะ', 'ลา', 'นอกเวลา', 'ว่าง']
             
    x = {}
    for p in all_pharmacists:
        for t in range(16):
            for task in tasks: x[p, t, task] = model.NewBoolVar(f'x_{p}_{t}_{task}')
            model.AddExactlyOne(x[p, t, task] for task in tasks)
            model.Add(x[p, t, 'ว่าง'] == 0)

    if DAY_OF_WEEK == 'Wed_Fri': break_slots, b_groups = [6, 7, 8, 9, 10, 11], [(6,8), (8,10), (10,12)] 
    else: break_slots, b_groups = [5, 6, 7, 8, 9, 10], [(5,7), (7,9), (9,11)]

    active_ft = []
    leave_slots = set()
    half_day_leaves = set() 
    
    for p in ft_pharmacists:
        if p in LEAVES:
            l_type = LEAVES[p]
            if l_type == 'ทั้งวัน': l_range = range(0, 16)
            elif l_type == 'เช้า': l_range = range(0, 9); half_day_leaves.add(p)
            elif l_type == 'บ่าย': l_range = range(7, 16); half_day_leaves.add(p)
            else: l_range = range(time_to_slot(l_type[0]), time_to_slot(l_type[1])); half_day_leaves.add(p)
            if l_type != 'ทั้งวัน': active_ft.append(p)
            for t in l_range: model.Add(x[p, t, 'ลา'] == 1); leave_slots.add((p, t))
        else: active_ft.append(p)

    for p in ft_pharmacists:
        for t in range(16):
            if (p, t) not in leave_slots: model.Add(x[p, t, 'ลา'] == 0)

    custom_dict_index = {}
    custom_task_slots_count = {p: 0 for p in ft_pharmacists} 
    for (p, start, end), task_name in CUSTOM_TASKS.items():
        s_idx, e_idx = time_to_slot(start), time_to_slot(end)
        for t in range(s_idx, e_idx):
            if (p, t) not in leave_slots:
                model.Add(x[p, t, 'งานเฉพาะ'] == 1)
                custom_dict_index[(p, t)] = task_name
                if p in ft_pharmacists: custom_task_slots_count[p] += 1
            
    for p in all_pharmacists:
        for t in range(16):
            if (p, t) not in custom_dict_index: model.Add(x[p, t, 'งานเฉพาะ'] == 0)

    for p in SICK_PEOPLE:
        if p in all_pharmacists:
            for t in range(16):
                for task in dispensing_tasks: model.Add(x[p, t, task] == 0)

    for (p, start, end), task_name in FIXED_MAIN_TASKS.items():
        s_idx, e_idx = time_to_slot(start), time_to_slot(end)
        for t in range(s_idx, e_idx): 
            if (p, t) not in leave_slots: model.Add(x[p, t, task_name] == 1)

    reward_vars = []

    for p in head_pharmacists:
        fixed_slots = set()
        for (fp, s, e), t_name in FIXED_MAIN_TASKS.items():
            if fp == p:
                for t in range(time_to_slot(s), time_to_slot(e)): fixed_slots.add(t)
        for t in range(16):
            if t not in fixed_slots:
                for task in dispensing_tasks + ver_cpoe_tasks + ver_ps_tasks + ['Match + C', 'Match + C2']:
                    if ALLOW_HEAD_ASSIST: reward_vars.append(x[p, t, task] * -5000000)
                    else: model.Add(x[p, t, task] == 0)

    for pt in PART_TIME:
        p = pt['name']
        s_idx, e_idx = time_to_slot(pt['start']), time_to_slot(pt['end'])
        my_dispense_allowed = ['จ่าย 7', 'จ่าย 8']
        if len(PART_TIME) > 2: my_dispense_allowed.extend(['จ่าย 6', 'จ่าย 9'])
        pt_all_allowed = my_dispense_allowed + ['Matching', 'พัก', 'นอกเวลา']
        
        for t in range(16): 
            if t < s_idx or t >= e_idx: model.Add(x[p, t, 'นอกเวลา'] == 1)
            else: model.Add(x[p, t, 'นอกเวลา'] == 0)
        
        for t in range(max(0, s_idx), min(16, e_idx)): model.Add(sum(x[p, t, task] for task in pt_all_allowed) == 1)

        b_type, b_time = pt.get('break_type', 'ไม่พักเลย'), pt.get('break_time', None)
        if b_type != "ไม่พักเลย" and b_time:
            b_s_idx = time_to_slot(b_time)
            if b_type == "พัก 1 ชั่วโมง":
                if b_s_idx < 16: model.Add(x[p, b_s_idx, 'พัก'] == 1)
                if b_s_idx + 1 < 16: model.Add(x[p, b_s_idx + 1, 'พัก'] == 1)
                for t in range(16):
                    if t != b_s_idx and t != (b_s_idx + 1): model.Add(x[p, t, 'พัก'] == 0)
            elif b_type == "พักครึ่งชั่วโมง":
                if b_s_idx < 16: model.Add(x[p, b_s_idx, 'พัก'] == 1)
                for t in range(16):
                    if t != b_s_idx: model.Add(x[p, t, 'พัก'] == 0)
        else:
            for t in range(16): model.Add(x[p, t, 'พัก'] == 0)

        for d in my_dispense_allowed:
            d_sum = sum(x[p, t, d] for t in range(16))
            over_2 = model.NewIntVar(0, 16, f'pt_over_2_{p}_{d}')
            model.Add(over_2 >= d_sum - 2); model.Add(over_2 >= 0)
            reward_vars.append(over_2 * -80000) 

        d7_sum = sum(x[p, t, 'จ่าย 7'] for t in range(16))
        d8_sum = sum(x[p, t, 'จ่าย 8'] for t in range(16))
        diff_78 = model.NewIntVar(-16, 16, f'diff_78_{p}')
        model.Add(diff_78 == d7_sum - d8_sum)
        abs_diff_78 = model.NewIntVar(0, 16, f'abs_diff_78_{p}')
        model.AddAbsEquality(abs_diff_78, diff_78)
        reward_vars.append(abs_diff_78 * -30000)
        
    for p in pt_pharmacists:
        for t in range(14): model.Add(sum(x[p, t+k, 'Matching'] for k in range(3)) <= 2)

    b_group_vars_ft = {0: [], 1: [], 2: []}
    full_day_active_ft = [p for p in active_ft if p not in half_day_leaves]
    normal_ft_for_break = [p for p in full_day_active_ft if p not in head_pharmacists]
    
    for p in all_pharmacists:
        if p in ft_pharmacists:
            model.Add(sum(x[p, t, 'นอกเวลา'] for t in range(16)) == 0) 
            if p not in head_pharmacists:
                for t in range(16): model.Add(x[p, t, 'Matching'] == 0) 
        
        if p in full_day_active_ft:
            if p in head_pharmacists:
                break_sum = sum(x[p, t, 'พัก'] for t in range(16))
                model.Add(break_sum <= 2)
                reward_vars.append(break_sum * 100000) 
                for t in range(12, 16): model.Add(x[p, t, 'พัก'] == 0)
                for t in range(0, 5): model.Add(x[p, t, 'พัก'] == 0)
                head_is_busy = False
                for t in break_slots:
                    if (p, t) in custom_slots_check or (p, t) in fixed_slots_check: head_is_busy = True
                if not head_is_busy:
                    for bg_idx, bg_range in enumerate(b_groups):
                        is_in_this_break = model.NewBoolVar(f'head_{p}_in_break_{bg_idx}')
                        model.Add(sum(x[p, t, 'พัก'] for t in range(*bg_range)) == 2).OnlyEnforceIf(is_in_this_break)
                        reward_vars.append(is_in_this_break * 200000)
            else:
                model.Add(sum(x[p, t, 'พัก'] for t in range(16)) == 2)
                choices = [model.NewBoolVar(f'choice_{p}_b{i}') for i in range(3)]
                if p in FIX_BREAKS and p in ft_pharmacists:
                    req_b = FIX_BREAKS[p]
                    for i in range(3): model.Add(choices[i] == (1 if i == req_b else 0))
                else:
                    model.AddExactlyOne(choices) 
                for i in range(3):
                    b_group_vars_ft[i].append(choices[i])
                    for t in range(*b_groups[i]): model.Add(x[p, t, 'พัก'] == 1).OnlyEnforceIf(choices[i])
                for t in range(16):
                    if t not in break_slots: model.Add(x[p, t, 'พัก'] == 0)
        elif p in ft_pharmacists:
            if p in head_pharmacists:
                break_sum = sum(x[p, t, 'พัก'] for t in range(16))
                model.Add(break_sum <= 2)
                reward_vars.append(break_sum * 100000)
                for t in range(12, 16): model.Add(x[p, t, 'พัก'] == 0)
                for t in range(0, 5): model.Add(x[p, t, 'พัก'] == 0)
            else:
                model.Add(sum(x[p, t, 'พัก'] for t in range(16)) == 0)

    total_active_ft_break = len(normal_ft_for_break) 
    if total_active_ft_break > 0:
        max_b_per_group = (total_active_ft_break // 3) + 2
        for i in range(3):
            model.Add(sum(b_group_vars_ft[i]) <= max_b_per_group) 
            model.Add(sum(b_group_vars_ft[i]) >= max(0, (total_active_ft_break // 3) - 1))

    for t in range(16):
        for task in tasks:
            if task not in ['พัก', 'งานเฉพาะ', 'ลา', 'นอกเวลา', 'ว่าง', 'Matching', 'Match + C2']:
                model.Add(sum(x[p, t, task] for p in all_pharmacists) <= 1)
        model.Add(sum(x[p, t, 'Match + C2'] for p in all_pharmacists) <= 1)

        if t < 2: 
            req_core = ['จ่าย 6', 'จ่าย 7', 'จ่าย 8', 'Ver 1 INC', 'Ver 2/ปณ.', 'Ver 3/A', 'Ver PS1', 'Match + C']
            reward_vars.append(sum(x[p, t, 'จ่าย 9'] for p in all_pharmacists) * 50000)
            model.Add(sum(x[p, t, 'จ่าย 9'] for p in all_pharmacists) <= 1)
        elif t == 2: req_core = ['จ่าย 5', 'จ่าย 6', 'จ่าย 7', 'จ่าย 8', 'จ่าย 9', 'Ver 1 INC', 'Ver 2/ปณ.', 'Ver 3/A', 'Ver PS1', 'Match + C']
        else: req_core = ['จ่าย 5', 'จ่าย 6', 'จ่าย 7', 'จ่าย 8', 'จ่าย 9', 'จ่าย 10', 'Ver 1 INC', 'Ver 2/ปณ.', 'Ver 3/A', 'Ver PS1', 'Match + C']
            
        for task in req_core: model.Add(sum(x[p, t, task] for p in all_pharmacists) == 1)

        if t < 2:
            model.Add(sum(x[p, t, 'จ่าย 4'] for p in all_pharmacists) == 0)
            model.Add(sum(x[p, t, 'จ่าย 5'] for p in all_pharmacists) == 0)
            model.Add(sum(x[p, t, 'จ่าย 10'] for p in all_pharmacists) == 0)
            model.Add(sum(x[p, t, 'จ่าย 11'] for p in all_pharmacists) == 0)

        if t not in break_slots: model.Add(sum(x[p, t, 'Ver PS2'] for p in all_pharmacists) == 1)
        else:
            model.Add(sum(x[p, t, 'Ver PS2'] for p in all_pharmacists) <= 1)
            ps2_sum = sum(x[p, t, 'Ver PS2'] for p in all_pharmacists)
            reward_vars.append(ps2_sum * 100000)

        if t < 3:
            model.Add(sum(x[p, t, 'จ่าย 10'] for p in all_pharmacists) <= 1)
            d10_sum = sum(x[p, t, 'จ่าย 10'] for p in all_pharmacists)
            reward_vars.append(d10_sum * 150000)

    for t in range(16):
        for i in range(2, 10): model.Add(sum(x[p, t, f'Ver PS{i+1}'] for p in all_pharmacists) <= sum(x[p, t, f'Ver PS{i}'] for p in all_pharmacists))
        for i in range(4, 10): model.Add(sum(x[p, t, f'Ver {i+1}'] for p in all_pharmacists) <= sum(x[p, t, f'Ver {i}'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่าย 8'] for p in all_pharmacists) <= sum(x[p, t, 'จ่าย 7'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่าย 6'] for p in all_pharmacists) <= sum(x[p, t, 'จ่าย 8'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่าย 9'] for p in all_pharmacists) <= sum(x[p, t, 'จ่าย 6'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่าย 5'] for p in all_pharmacists) <= sum(x[p, t, 'จ่าย 9'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่าย 10'] for p in all_pharmacists) <= sum(x[p, t, 'จ่าย 5'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่าย 4'] for p in all_pharmacists) <= sum(x[p, t, 'จ่าย 10'] for p in all_pharmacists))
        model.Add(sum(x[p, t, 'จ่าย 11'] for p in all_pharmacists) <= sum(x[p, t, 'จ่าย 4'] for p in all_pharmacists))

    for p in all_pharmacists:
        for t in range(15):
            for task1 in dispensing_tasks:
                for task2 in dispensing_tasks:
                    if task1 != task2: model.AddImplication(x[p, t, task1], x[p, t+1, task2].Not())

    for p in all_pharmacists:
        for cat in [dispensing_tasks, ver_cpoe_tasks, ver_ps_tasks, ['Match + C', 'Match + C2']]:
            for t in range(14): model.Add(sum(x[p, t+k, task] for task in cat for k in range(3)) <= 2)

    is_disp_7_vars = []
    for p in ft_pharmacists:
        if p not in head_pharmacists:
            tot_disp = sum(x[p, t, task] for t in range(16) for task in dispensing_tasks)
            over_3hr_var = model.NewBoolVar(f'over_3hr_{p}')
            model.Add(tot_disp <= 6 + over_3hr_var)
            model.Add(tot_disp <= 7) 
            reward_vars.append(over_3hr_var * -500000) 
            is_disp_7_vars.append(over_3hr_var)
            if p in active_ft and p not in SICK_PEOPLE:
                has_heavy_custom_tasks = custom_task_slots_count[p] >= 6 
                is_half_day_leave = p in half_day_leaves
                short_disp = model.NewIntVar(0, 16, f'short_disp_{p}')
                if has_heavy_custom_tasks or is_half_day_leave: model.Add(short_disp >= 2 - tot_disp)
                else: model.Add(short_disp >= 4 - tot_disp)
                model.Add(short_disp >= 0)
                reward_vars.append(short_disp * -500000) 

            for d in ['จ่าย 6', 'จ่าย 7', 'จ่าย 8', 'จ่าย 9']: model.Add(sum(x[p, t, d] for t in range(16)) <= 2)
            for d in ['จ่าย 4', 'จ่าย 5', 'จ่าย 10', 'จ่าย 11']:
                total_d = sum(x[p, t, d] for t in range(16))
                over_d = model.NewIntVar(0, 16, f'over_{p}_{d}')
                model.Add(over_d >= total_d - 2); model.Add(over_d >= 0) 
                reward_vars.append(over_d * -2500) 

            done_disp_7 = model.NewBoolVar(f'done_disp_7_{p}')
            model.Add(sum(x[p, t, 'จ่าย 7'] for t in range(16)) > 0).OnlyEnforceIf(done_disp_7)
            model.Add(sum(x[p, t, 'จ่าย 7'] for t in range(16)) == 0).OnlyEnforceIf(done_disp_7.Not())
            done_disp_8 = model.NewBoolVar(f'done_disp_8_{p}')
            model.Add(sum(x[p, t, 'จ่าย 8'] for t in range(16)) > 0).OnlyEnforceIf(done_disp_8)
            model.Add(sum(x[p, t, 'จ่าย 8'] for t in range(16)) == 0).OnlyEnforceIf(done_disp_8.Not())
            model.Add(done_disp_7 + done_disp_8 <= 1)
            model.Add(sum(x[p, t, 'Match + C'] + x[p, t, 'Match + C2'] for t in range(16)) <= 2)

    model.Add(sum(is_disp_7_vars) <= 2) 

    for p in all_pharmacists:
        for t in range(14):
            is_disp_t = sum(x[p, t, d] for d in dispensing_tasks)
            is_disp_t1 = sum(x[p, t+1, d] for d in dispensing_tasks)
            is_disp_t2 = sum(x[p, t+2, d] for d in dispensing_tasks)
            too_long = model.NewBoolVar(f'too_long_disp_{p}_{t}')
            model.Add(is_disp_t + is_disp_t1 + is_disp_t2 <= 2 + too_long)
            reward_vars.append(too_long * -100000) 
            short_break = model.NewBoolVar(f'short_break_disp_{p}_{t}')
            model.Add(is_disp_t - is_disp_t1 + is_disp_t2 <= 1 + short_break)
            if p in ft_pharmacists: reward_vars.append(short_break * -2000000) 
            else: reward_vars.append(short_break * -500000)

    for p in all_pharmacists:
        for t in range(15):
            for task in dispensing_tasks + ver_cpoe_tasks + ver_ps_tasks + ['Match + C', 'Match + C2']:
                match_var = model.NewBoolVar(f'pair_{p}_{t}_{task}')
                model.AddImplication(match_var, x[p, t, task])
                model.AddImplication(match_var, x[p, t+1, task])
                if task in dispensing_tasks: reward_vars.append(match_var * 500000) 
                else: reward_vars.append(match_var * 150000)

    for p in ft_pharmacists:
        ft_iso_disp_vars = []
        for t in range(16):
            for d in dispensing_tasks:
                iso_disp = model.NewBoolVar(f'iso_disp_{p}_{t}_{d}')
                prev_v = x[p, t-1, d] if t > 0 else 0
                next_v = x[p, t+1, d] if t < 15 else 0
                model.Add(x[p, t, d] - prev_v - next_v <= iso_disp)
                ft_iso_disp_vars.append(iso_disp)
                reward_vars.append(iso_disp * -200000) 
        model.Add(sum(ft_iso_disp_vars) <= 2)

    for p in all_pharmacists:
        for t in range(16):
            for target_task in ['Ver 1 INC', 'Ver 2/ปณ.', 'Ver 3/A', 'Match + C', 'Ver PS1']:
                iso_var = model.NewBoolVar(f'iso_{target_task}_{p}_{t}')
                prev_v = x[p, t-1, target_task] if t > 0 else 0
                next_v = x[p, t+1, target_task] if t < 15 else 0
                model.Add(x[p, t, target_task] - prev_v - next_v <= iso_var)
                reward_vars.append(iso_var * -100000)

    for t in range(16):
        weights = {
            'จ่าย 7': 400000, 'จ่าย 8': 390000, 'จ่าย 6': 380000, 'จ่าย 9': 370000,
            'จ่าย 5': 360000, 'จ่าย 10': 350000, 'จ่าย 4': 300000, 'จ่าย 11': 290000, 
            'Ver 4': 50000, 'Ver PS3': 48000, 
            'Match + C2': 47000, 
            'Ver 5': 46000, 'Ver PS4': 44000, 
            'Ver 6': 42000, 'Ver PS5': 40000, 
            'Ver 7': 38000, 'Ver PS6': 36000, 
            'Ver 8': 34000, 'Ver PS7': 32000, 
            'Ver 9': 30000, 'Ver PS8': 28000, 
            'Ver 10': 26000, 'Ver PS9': 24000, 'Ver PS10': 22000,
            'Ver 1 INC': 85000, 'Ver 2/ปณ.': 80000, 'Ver 3/A': 75000, 
            'Ver PS1': 70000, 'Ver PS2': 65000, 'Match + C': 60000, 'Matching': 20000
        }
        if IS_MWF and (t in break_slots):
            weights['จ่าย 4'] = -50000 
            weights['จ่าย 11'] = -50000

        for task, weight in weights.items():
            for i, p in enumerate(all_pharmacists): reward_vars.append(x[p, t, task] * (weight + i))

    model.Maximize(sum(reward_vars))
    solver = cp_model.CpSolver()
    solver.parameters.num_workers = 8
    solver.parameters.max_time_in_seconds = 20.0  

    status = solver.Solve(model)

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        schedule_data = []
        for p in all_pharmacists:
            row_data = {'ชื่อ/เวลา': p} 
            for t in range(16):
                assigned = ""
                for tsk in tasks:
                    if solver.Value(x[p, t, tsk]) == 1:
                        if tsk == 'งานเฉพาะ': assigned = custom_dict_index.get((p, t), 'งานเฉพาะ')
                        elif tsk in ['นอกเวลา', 'ว่าง']: assigned = '-'
                        elif tsk == 'ลา': assigned = 'ลา'
                        else:
                            if ALLOW_HEAD_ASSIST and p in head_pharmacists and tsk in base_main_tasks:
                                assigned = f"{tsk} (ช่วย)"
                            else: assigned = tsk
                row_data[time_slots[t]] = assigned
            schedule_data.append(row_data)
            
        df_result = pd.DataFrame(schedule_data)
        summary_row = {'ชื่อ/เวลา': 'P/C/D'} 
        for t_idx in range(16):
            time_col = time_slots[t_idx]
            w_count, x_count, disp_nums = 0, 0, []
            for p in all_pharmacists:
                val_str = str(df_result.loc[df_result['ชื่อ/เวลา'] == p, time_col].values[0])
                if 'Ver PS' in val_str: w_count += 1
                elif 'Ver' in val_str: x_count += 1
                elif 'จ่าย ' in val_str:
                    try: disp_nums.append(int(val_str.replace('จ่าย ', '').replace('(ช่วย)', '').strip()))
                    except: pass
            yz_str = f"{min(disp_nums)}-{max(disp_nums)}" if disp_nums else "-"
            summary_row[time_col] = f"{w_count}/{x_count}/{yz_str}"
        df_result = pd.concat([df_result, pd.DataFrame([summary_row])], ignore_index=True)
        
        return df_result, "Success", ""
    else: return None, "Infeasible", "เงื่อนไขตึงเกินไป หรือคนไม่พอจัดตาราง"

# ------------------------------------------------------------------
# UI Rendering functions
# ------------------------------------------------------------------
def get_cell_bg_hex(val_str):
    val = str(val_str)
    if '/' in val and '-' in val and val and val[0].isdigit(): return "FFF2CC"
    elif 'จ่าย ' in val: return "D5E8D4"
    elif 'Match' in val: return "DAE8FC"
    elif val == 'Matching': return "DAE8FC"
    elif 'Ver PS' in val: return "E1D5E7"
    elif 'Ver' in val: return "FFE6CC"
    elif 'พัก' in val: return "F8CECC"
    elif val == 'ลา': return "E6E6E6" 
    elif val in ['-', 'ว่าง', 'นอกเวลา']: return "F5F5F5"
    return "F5F5F5"

def get_header_color(t_idx, day_of_week):
    if day_of_week == 'Normal':
        if t_idx in [0, 1, 3, 4, 11, 12]: return 'orange' 
        if t_idx in [2]: return 'yellow'                 
        if t_idx in [5, 6, 9, 10]: return 'pink'         
        if t_idx in [7, 8]: return 'purple'              
        if t_idx in [13, 14, 15]: return 'blue'          
    else: 
        if t_idx in [0, 1, 4, 5, 12, 13]: return 'orange' 
        if t_idx in [2, 3]: return 'yellow'              
        if t_idx in [6, 7, 10, 11]: return 'pink'        
        if t_idx in [8, 9]: return 'purple'              
        if t_idx in [14, 15]: return 'blue'              
    return None

header_color_map = {
    'orange': PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid'),
    'yellow': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    'pink': PatternFill(start_color='F8CECC', end_color='F8CECC', fill_type='solid'),
    'purple': PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid'),
    'blue': PatternFill(start_color='DAE8FC', end_color='DAE8FC', fill_type='solid')
}
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def build_html_table(df, selected_date_str, DAY_OF_WEEK):
    thai_date_str = get_thai_date(selected_date_str)
    def get_cell_style(val_str):
        bg = "#" + get_cell_bg_hex(val_str)
        color, weight = "black", "normal"
        if '/' in val_str and '-' in val_str and val_str and val_str[0].isdigit(): weight = "bold"
        elif 'Match' in val_str and ('C' in val_str or 'C2' in val_str): color, weight = "red", "bold"
        elif val_str == 'ลา': color, weight = "black", "normal" 
        elif val_str in ['-', 'ว่าง', 'นอกเวลา']: color = "#808080"
        return f"background-color: {bg}; color: {color}; font-weight: {weight}; border: 1px solid black; padding: 2px; text-align: center; font-size: 14px; white-space: nowrap; height: 30px; box-sizing: border-box;"
        
    def get_head_color_hex(t_idx, day_of_week):
        color_name = get_header_color(t_idx, day_of_week)
        mapping = {'orange': '#FFE6CC', 'yellow': '#FFF2CC', 'pink': '#F8CECC', 'purple': '#E1D5E7', 'blue': '#DAE8FC'}
        return mapping.get(color_name, '#FFFFFF')

    cols = df.columns.tolist()
    num_cols = len(cols)
    
    html = f"""
    <div id='capture-area' style='background-color: white; padding: 15px; display: inline-block; font-family: "Sarabun", sans-serif;'>
        <table style='border-collapse: collapse; table-layout: fixed; width: max-content; border: 2px solid #ddd; box-shadow: 0 4px 12px rgba(0,0,0,0.05); border-radius: 8px; overflow: hidden;'>
            <tr><td colspan='{num_cols}' style='text-align: center; font-size: 24px; font-weight: bold; border: none; padding-top: 5px; color: #333;'>ตารางปฏิบัติงานเภสัชกร ห้องยาชั้น 1 อาคารสมเด็จพระเทพรัตน์</td></tr>
            <tr><td colspan='{num_cols}' style='text-align: center; font-size: 18px; font-weight: normal; border: none; padding-bottom: 15px; color: #666;'>ประจำ{thai_date_str}</td></tr>
            <tr>
    """
    for i, col in enumerate(cols):
        bg = "#FFFFFF" if i == 0 else get_head_color_hex(i - 1, DAY_OF_WEEK)
        col_width = "69px"
        html += f"<th style='background-color: {bg}; color: #333; border: 1px solid #ddd; padding: 2px; font-size: 15px; font-weight: bold; white-space: nowrap; height: 40px; width: {col_width}; min-width: {col_width}; max-width: {col_width}; overflow: hidden;'>{col}</th>"
    html += "</tr>"
    
    for _, row in df.iterrows():
        html += "<tr style='height: 30px;'>"
        for i, col in enumerate(cols):
            val = row[col]
            style = get_cell_style(val)
            col_width = "69px"
            if i == 0: style = "background-color: #FFFFFF; color: #333; font-weight: bold; border: 1px solid #ddd; padding: 2px; text-align: center; font-size: 15px;"
            if _ == len(df)-1: style = style.replace("font-weight: normal", "font-weight: bold")
            html += f"<td style='{style} width: {col_width}; min-width: {col_width}; max-width: {col_width}; overflow: hidden;'>{val}</td>"
        html += "</tr>"
    html += "</table></div>"
    return html

# ==================================================================
# หน้า 5: จัดการผู้ใช้งาน
# ==================================================================
if page == "👥 จัดการผู้ใช้งาน":
    st.title("👥 จัดการรายชื่อและสิทธิ์แอปพลิเคชัน")
    
    with st.form("add_user_form", clear_on_submit=True):
        c1, c2, c3, c4 = st.columns(4)
        with c1: new_user = st.text_input("Username (ใช้ล็อกอิน)")
        with c2: new_pass = st.text_input("Password", type="password")
        with c3: new_name = st.text_input("ชื่อเล่น (แสดงในตาราง AI)")
        with c4: new_role = st.selectbox("สิทธิ์ (Role)", ["Staff", "Head", "Admin"])
        
        c5, c6, c7, c8 = st.columns(4)
        with c5: real_name = st.text_input("ชื่อจริง (Real Name)")
        with c6: surname = st.text_input("นามสกุล (Surname)")
        with c7: email = st.text_input("อีเมล (ใช้กู้รหัส)")
        with c8: position = st.selectbox("ตำแหน่งงาน", ["เภสัชกร", "ผู้ช่วยเภสัชกร"], index=0 if current_view=="เภสัชกร" else 1)
        
        c9, c10, c11, c12 = st.columns(4)
        with c9: display_order_str = st.text_input("ลำดับในตาราง (1,2,3...)", value="99")
        
        if st.form_submit_button("บันทึกพนักงานใหม่", type="primary") and new_user and new_name:
            try: display_order = int(display_order_str)
            except ValueError: display_order = 99
            add_user_db(new_user, new_pass, new_name, new_role, real_name, surname, email, position, display_order)
            st.toast("✅ เพิ่มข้อมูลสำเร็จ!"); time.sleep(1); st.rerun()
            
    st.divider()
    st.subheader(f"พนักงานในระบบ (หมวด: {current_view})")
    
    sorted_users = sorted(users_db.values(), key=lambda x: (x.get('display_order') if x.get('display_order') is not None else 99, x.get('full_name', '')))
    for u in sorted_users:
        if u.get('role') == 'System' or get_std_pos(u) != current_view: continue
        with st.container(border=True):
            # แถวบน: Username, ชื่อในตาราง, ตำแหน่ง, ลำดับ
            c1, c2, c3, c4 = st.columns([1.5, 2.5, 2.5, 1.5])
            with c1:
                st.caption("Username")
                st.markdown(f"**{u['username']}**")
            with c2:
                new_fname = st.text_input("ชื่อในตาราง", value=u['full_name'], key=f"fname_{u['username']}")
                if new_fname != u['full_name'] and new_fname.strip() != "":
                    update_user_fullname(u['username'], new_fname.strip())
                    st.toast("✅ อัปเดตชื่อในตารางสำเร็จ!")
                    time.sleep(0.5)
                    st.rerun()
            with c3:
                pos_opts = ["เภสัชกร", "ผู้ช่วยเภสัชกร"]
                curr_p_idx = safe_idx(pos_opts, get_std_pos(u), 0)
                new_p = st.selectbox("ตำแหน่งงาน", pos_opts, index=curr_p_idx, key=f"pos_{u['username']}")
                if new_p != get_std_pos(u):
                    update_user_profile(u['username'], u.get('real_name',''), u.get('surname',''), u.get('email',''), new_p)
                    st.toast("✅ ย้ายตำแหน่งพนักงานสำเร็จ!")
                    time.sleep(0.5)
                    st.rerun()
            with c4:
                curr_order = u.get('display_order') if u.get('display_order') is not None else 99
                new_ord_str = st.text_input("ลำดับในตาราง", value=str(curr_order), key=f"ord_{u['username']}")
                try: new_ord = int(new_ord_str)
                except ValueError: new_ord = curr_order
                if new_ord != curr_order:
                    update_user_order(u['username'], new_ord)
                    st.toast("✅ อัปเดตลำดับสำเร็จ!")
                    time.sleep(0.5)
                    st.rerun()
                    
            # แถวล่าง: ชื่อ-สกุลจริง, สิทธิ์, ลบ
            c1b, c2b, c3b, c4b = st.columns([1.5, 2.5, 2.5, 1.5])
            with c1b:
                real_n = u.get('real_name') or '-'
                sur_n = u.get('surname') or ''
                st.caption("ชื่อ-นามสกุล")
                st.markdown(f"<span style='color:#555;'>{real_n} {sur_n}</span>", unsafe_allow_html=True)
            with c2b:
                st.caption("อีเมล (Email)")
                st.markdown(f"<span style='color:#555;'>{u.get('email') or '-'}</span>", unsafe_allow_html=True)
            with c3b:
                role_opts = ["Staff", "Head", "Admin"]
                curr_r_idx = safe_idx(role_opts, u.get('role', 'Staff'), 0)
                new_r = st.selectbox("สิทธิ์ (Role)", role_opts, index=curr_r_idx, key=f"role_{u['username']}")
                if new_r != u['role']: 
                    update_user_role(u['username'], new_r)
                    st.rerun()
            with c4b:
                st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
                if u['username'] != user_info['username']: 
                    if st.button("🗑️ ลบผู้ใช้", key=f"del_u_{u['username']}", use_container_width=True):
                        delete_user_db(u['username'])
                        st.toast("✅ ลบข้อมูลสำเร็จ!")
                        time.sleep(1)
                        st.rerun()
